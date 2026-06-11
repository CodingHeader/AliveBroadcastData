from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session as DBSession
from sqlalchemy import func
from fastapi.responses import FileResponse, StreamingResponse
from database import get_db
from models import Setting, Deal, Lead, Anchor, SessionAnchor, Session, SchedulePlan, ScheduleSlot, ScheduleBinding, AdAccount, RoomAccountBinding, DashboardTab, ClueConfig, TabAnalysis, RecruitTeam, RecruitEmployee, ClueAssignment, ApiClue, AdPlan, AdPlanSpend, Comment, PrivateMessage
from auth import verify_password, create_token, get_current_admin, hash_password
from pydantic import BaseModel, field_validator
from typing import Optional
from datetime import datetime
from pathlib import Path
import json, os, shutil, io

router = APIRouter()

class LoginRequest(BaseModel): username: str; password: str
class DealCreate(BaseModel): session_id: Optional[int]=None; lead_id: Optional[int]=None; customer_name: str; amount: float; deal_time: str; employee: Optional[str]=None; notes: Optional[str]=None
class ManualClueCreate(BaseModel):
    """手动添加私信线索请求模型"""
    name: Optional[str] = None
    phone: Optional[str] = None
    weixin: Optional[str] = None
    message_time: Optional[str] = None  # 私信时间，用于自动匹配场次
    session_id: Optional[int] = None    # 直接指定场次ID（优先级高于message_time）
    notes: Optional[str] = None
    source: Optional[str] = "marketing"  # 来源：marketing(营销流), natural(自然流), manual(手动添加)

    @field_validator("session_id", mode="before")
    @classmethod
    def coerce_session_id(cls, v):
        if v == "" or v is None:
            return None
        try:
            return int(v)
        except (ValueError, TypeError):
            return None

    @field_validator("message_time", mode="before")
    @classmethod
    def coerce_message_time(cls, v):
        if v == "" or v is None:
            return None
        return v

@router.post("/login")
def login(req: LoginRequest, db: DBSession = Depends(get_db)):
    stored_username = db.query(Setting).filter(Setting.key == "admin_username").first()
    stored_password = db.query(Setting).filter(Setting.key == "admin_password").first()
    if not stored_username or req.username != stored_username.value: raise HTTPException(401, detail="账号或密码错误")
    if not verify_password(req.password, stored_password.value): raise HTTPException(401, detail="账号或密码错误")
    return {"code": 0, "token": create_token(req.username)}

@router.get("/settings")
def get_settings(admin=Depends(get_current_admin), db=Depends(get_db)):
    settings = db.query(Setting).all()
    result = {s.key: s.value for s in settings}
    for key in ['ai_api_key', 'email_password']:
        if key in result and result[key]: result[key] = result[key][:4] + '****'
    return {"code":0, "data": result}

@router.put("/settings")
def update_settings(data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    for key, value in data.items():
        setting = db.query(Setting).filter(Setting.key == key).first()
        if setting: setting.value = str(value); setting.updated_at = datetime.now().isoformat()
        else: db.add(Setting(key=key, value=str(value)))
    db.commit(); return {"code":0}

@router.get("/deals")
def list_deals(page: int = 1, size: int = 20, admin=Depends(get_current_admin), db=Depends(get_db)):
    query = db.query(Deal).order_by(Deal.created_at.desc())
    total = query.count(); items = query.offset((page-1)*size).limit(size).all()
    return {"code":0, "data":{"items":[{"id":d.id,"customer_name":d.customer_name,"amount":d.amount,"deal_time":d.deal_time,"employee":d.employee,"lead_nickname":d.lead.nickname if d.lead else None,"lead_id":d.lead_id,"session_id":d.session_id} for d in items], "total":total}}

@router.post("/deals")
def create_deal(data: DealCreate, admin=Depends(get_current_admin), db=Depends(get_db)):
    if data.lead_id:
        lead = db.query(Lead).get(data.lead_id)
        if lead and lead.is_deal:
            return {"code": 400, "message": "该线索已标记成单"}
    deal = Deal(**data.model_dump()); db.add(deal)
    if data.lead_id:
        lead = db.query(Lead).get(data.lead_id)
        if lead: lead.is_deal = True
    db.commit(); return {"code":0, "deal_id": deal.id}

@router.delete("/deals/{id}")
def delete_deal(id: int, admin=Depends(get_current_admin), db=Depends(get_db)):
    deal = db.query(Deal).get(id)
    if not deal: raise HTTPException(404)
    if deal.lead_id:
        other = db.query(Deal).filter(Deal.lead_id == deal.lead_id, Deal.id != id).count()
        if other == 0:
            lead = db.query(Lead).get(deal.lead_id)
            if lead: lead.is_deal = False
    db.delete(deal); db.commit(); return {"code":0}

@router.delete("/deals/by-lead/{lead_id}")
def delete_deal_by_lead(lead_id: int, admin=Depends(get_current_admin), db=Depends(get_db)):
    """按 lead_id 取消成单"""
    deal = db.query(Deal).filter(Deal.lead_id == lead_id).first()
    if not deal: return {"code": 404, "message": "未找到成单记录"}
    if deal.lead_id:
        lead = db.query(Lead).get(deal.lead_id)
        if lead: lead.is_deal = False
    db.delete(deal); db.commit(); return {"code":0}

@router.get("/anchors")
def list_anchors(admin=Depends(get_current_admin), db=Depends(get_db)):
    anchors = db.query(Anchor).all()
    return {"code":0, "data":[{"id":a.id,"name":a.name,"gender":a.gender,"style":a.style,"is_parttime":a.is_parttime or 0} for a in anchors]}

@router.post("/anchors")
def create_anchor(data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    try:
        anchor = Anchor(name=data["name"], gender=data.get("gender"), style=data.get("style"), is_parttime=data.get("is_parttime", 0))
        db.add(anchor); db.commit()
        return {"code":0, "anchor_id": anchor.id}
    except Exception as e:
        db.rollback()
        if "UNIQUE constraint" in str(e):
            return {"code":400, "message":"主播名称已存在"}
        return {"code":500, "message":f"添加失败: {str(e)}"}

@router.put("/anchors/{id}")
def update_anchor(id: int, data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    anchor = db.query(Anchor).get(id)
    if not anchor: raise HTTPException(404)
    try:
        if "name" in data: anchor.name = data["name"]
        if "gender" in data: anchor.gender = data["gender"]
        if "style" in data: anchor.style = data["style"]
        if "is_parttime" in data: anchor.is_parttime = data["is_parttime"]
        db.commit(); return {"code":0}
    except Exception as e:
        db.rollback()
        if "UNIQUE constraint" in str(e):
            return {"code":400, "message":"主播名称已存在"}
        return {"code":500, "message":f"保存失败: {str(e)}"}

@router.delete("/anchors/{id}")
def delete_anchor(id: int, admin=Depends(get_current_admin), db=Depends(get_db)):
    db.query(SessionAnchor).filter(SessionAnchor.anchor_id == id).delete()
    db.query(Anchor).filter(Anchor.id == id).delete(); db.commit(); return {"code":0}

@router.post("/sessions/{session_id}/anchors")
def bind_anchors(session_id: int, data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    db.query(SessionAnchor).filter(SessionAnchor.session_id == session_id).delete()
    for aid in data.get("anchor_ids", []): db.add(SessionAnchor(session_id=session_id, anchor_id=aid))
    db.commit(); return {"code":0}

@router.post("/email/test")
def test_email(admin=Depends(get_current_admin), db=Depends(get_db)):
    from services.email_service import send_email, get_email_config
    try:
        config = get_email_config(db)
        if not config.get("email_receivers"):
            return {"code": 400, "message": "请先添加收件人"}
        html = f"""<html><body style="font-family:Arial,sans-serif;padding:20px;">
            <h2 style="color:#2563eb;">AliveBroadcastData - SMTP Test</h2>
            <p>邮件配置测试成功！</p>
            <p>发送时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <hr style="border:1px solid #e5e7eb;margin:20px 0;">
            <p style="color:#6b7280;font-size:12px;">AliveBroadcastData System</p>
        </body></html>"""
        send_email("【AliveBroadcastData】邮件配置测试", html, config, config["email_receivers"])
        return {"code": 0, "message": "测试邮件已发送"}
    except Exception as e:
        return {"code": 500, "message": f"发送失败: {str(e)}"}

@router.put("/settings/password")
def change_password(data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    stored = db.query(Setting).filter(Setting.key == "admin_password").first()
    if not verify_password(data.get("old_password",""), stored.value): return {"code":400,"message":"原密码错误"}
    stored.value = hash_password(data["new_password"]); db.commit(); return {"code":0,"message":"密码修改成功"}

@router.get("/employee/stats")
def employee_stats(date_from: str = None, date_to: str = None, admin=Depends(get_current_admin), db=Depends(get_db)):
    from sqlalchemy import func
    query = db.query(Deal.employee, func.count(Deal.id).label('deal_count'), func.sum(Deal.amount).label('total_amount'))
    if date_from: query = query.filter(Deal.deal_time >= date_from)
    if date_to: query = query.filter(Deal.deal_time <= date_to)
    results = query.group_by(Deal.employee).all()
    return {"code":0, "data":[{"employee":r[0],"deal_count":r[1],"total_amount":float(r[2] or 0)} for r in results]}

@router.get("/anchor-income-stats")
def anchor_income_stats(date_from: str = None, date_to: str = None, anchor_id: int = None, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """主播收入统计API"""
    from sqlalchemy import case
    
    query = db.query(
        Session.id.label('session_id'),
        Session.start_time.label('session_start'),
        Session.end_time.label('session_end'),
        Session.duration_minutes,
        Anchor.id.label('anchor_id'),
        Anchor.name.label('anchor_name'),
        SessionAnchor.on_time,
        SessionAnchor.off_time,
        SessionAnchor.anchor_order,
    ).join(
        SessionAnchor, Session.id == SessionAnchor.session_id
    ).join(
        Anchor, SessionAnchor.anchor_id == Anchor.id
    )

    if date_from:
        query = query.filter(Session.start_time >= date_from)
    if date_to:
        query = query.filter(Session.start_time <= date_to + " 23:59:59")
    if anchor_id:
        query = query.filter(Anchor.id == anchor_id)

    rows = query.order_by(Session.start_time, SessionAnchor.anchor_order).all()

    session_counter = {}
    result = []
    for row in rows:
        date_str = row.session_start[:10] if row.session_start else ""

        if date_str not in session_counter:
            session_counter[date_str] = 0
        session_counter[date_str] += 1
        session_order = session_counter[date_str]

        on_time = row.on_time
        off_time = row.off_time

        if on_time and off_time:
            try:
                on_parts = on_time.split(":")
                off_parts = off_time.split(":")
                on_minutes = int(on_parts[0]) * 60 + int(on_parts[1])
                off_minutes = int(off_parts[0]) * 60 + int(off_parts[1])
                if off_minutes < on_minutes:
                    off_minutes += 24 * 60
                duration_hours = round((off_minutes - on_minutes) / 60.0, 2)
            except:
                duration_hours = 0
        else:
            duration_hours = round((row.duration_minutes or 0) / 60.0, 2)
            if not on_time:
                on_time = row.session_start[11:16] if row.session_start else ""
            if not off_time:
                off_time = row.session_end[11:16] if row.session_end else ""

        income = round(duration_hours * 0, 2)  # 不计算收入（排班未绑定主播）

        try:
            hour = int(on_time.split(":")[0])
            if 6 <= hour < 12:
                session_period = "上午场"
            elif 12 <= hour < 18:
                session_period = "下午场"
            else:
                session_period = "晚间场"
        except:
            session_period = "未知"

        lead_stats = db.query(
            func.count(Lead.id).label('total_leads'),
            func.sum(case((Lead.ad_account != None, 1), else_=0)).label('ad_leads'),
            func.sum(case(((Lead.ad_account == None) | (Lead.ad_account == '--'), 1), else_=0)).label('natural_leads')
        ).filter(Lead.session_id == row.session_id).first()

        comment_stats = db.query(
            func.count(Comment.id).label('comment_count'),
            func.sum(case((Comment.has_lead == True, 1), else_=0)).label('comment_leads')
        ).filter(Comment.session_id == row.session_id).first()

        pm_stats = db.query(
            func.count(PrivateMessage.id).label('pm_count'),
            func.sum(case((PrivateMessage.has_lead == True, 1), else_=0)).label('pm_lead_count')
        ).filter(PrivateMessage.session_id == row.session_id).first()

        result.append({
            "date": date_str,
            "anchor_name": row.anchor_name,
            "session_id": row.session_id,
            "session_order": session_order,
            "on_time": on_time,
            "off_time": off_time,
            "anchor_order": row.anchor_order,
            "session_period": session_period,
            "duration_hours": duration_hours,
            "income": income,
            "ad_lead_count": int(lead_stats.ad_leads or 0) if lead_stats else 0,
            "natural_lead_count": int(lead_stats.natural_leads or 0) if lead_stats else 0,
            "total_lead_count": int(lead_stats.total_leads or 0) if lead_stats else 0,
            "comment_count": comment_stats.comment_count if comment_stats else 0,
            "comment_lead_count": int(comment_stats.comment_leads or 0) if comment_stats else 0,
            "pm_count": pm_stats.pm_count if pm_stats else 0,
            "pm_lead_count": int(pm_stats.pm_lead_count or 0) if pm_stats else 0,
        })

    summary = {
        "total_income": sum(r["income"] for r in result),
        "total_hours": sum(r["duration_hours"] for r in result),
        "total_ad_leads": sum(r["ad_lead_count"] for r in result),
        "total_natural_leads": sum(r["natural_lead_count"] for r in result),
        "total_leads": sum(r["total_lead_count"] for r in result),
        "total_comments": sum(r["comment_count"] for r in result),
        "total_comment_leads": sum(r["comment_lead_count"] for r in result),
        "total_pm": sum(r["pm_count"] for r in result),
        "total_pm_leads": sum(r["pm_lead_count"] for r in result),
    }

    return {"code": 0, "data": result, "summary": summary}

@router.get("/anchor-salary-stats")
def anchor_salary_stats(date_from: str = None, date_to: str = None, anchor_id: int = None,
                         admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """主播月薪统计API - 按天+主播分组，拼接时段（兼容旧版）"""
    from services.anchor_stats_service import get_comprehensive_stats
    stats = get_comprehensive_stats(db, date_from, date_to, anchor_id)
    return {
        "code": 0,
        "summary": stats["anchor_summary"],
        "daily_detail": stats["daily_aggregated"],
    }

@router.get("/anchor-stats/comprehensive")
def anchor_comprehensive_stats(date_from: str = None, date_to: str = None,
                                anchor_id: int = None, anchor_name: str = None,
                                admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """主播综合分析统计API - 包含场次明细、关联场次、评论获客等全部字段"""
    from services.anchor_stats_service import get_comprehensive_stats, calc_anchor_salary, generate_excel_data

    # 获取参数
    base_rate = float(db.query(Setting).filter(Setting.key == "anchor_salary_base_rate").first().value if db.query(Setting).filter(Setting.key == "anchor_salary_base_rate").first() else 40)
    lead_commission = float(db.query(Setting).filter(Setting.key == "anchor_salary_lead_commission").first().value if db.query(Setting).filter(Setting.key == "anchor_salary_lead_commission").first() else 0)
    ad_commission = float(db.query(Setting).filter(Setting.key == "anchor_salary_ad_commission").first().value if db.query(Setting).filter(Setting.key == "anchor_salary_ad_commission").first() else 0)
    natural_commission = float(db.query(Setting).filter(Setting.key == "anchor_salary_natural_commission").first().value if db.query(Setting).filter(Setting.key == "anchor_salary_natural_commission").first() else 0)

    stats = get_comprehensive_stats(db, date_from, date_to, anchor_id, anchor_name)

    # 计算薪资
    salary_params = {
        "base_rate": base_rate,
        "lead_commission": lead_commission,
        "ad_commission": ad_commission,
        "natural_commission": natural_commission,
    }
    salary_results = calc_anchor_salary(**salary_params, anchor_summary=stats["anchor_summary"])

    # 生成Excel数据
    excel_data = generate_excel_data(stats["daily_detail"], stats["anchor_summary"], salary_params)

    return {
        "code": 0,
        "daily_detail": stats["daily_detail"],
        "daily_aggregated": stats["daily_aggregated"],
        "anchor_summary": salary_results,
        "overall_summary": stats["overall_summary"],
        "excel_data": excel_data,
        "salary_params": salary_params,
    }

@router.post("/anchor-stats/salary-calc")
def anchor_salary_calc(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """主播月薪计算API - 使用动态参数计算"""
    from services.anchor_stats_service import get_comprehensive_stats, calc_anchor_salary

    date_from = data.get("date_from")
    date_to = data.get("date_to")
    anchor_id = data.get("anchor_id")
    anchor_name = data.get("anchor_name")
    base_rate = data.get("base_rate", 40)
    lead_commission = data.get("lead_commission", 0)
    ad_commission = data.get("ad_commission", 0)
    natural_commission = data.get("natural_commission", 0)

    stats = get_comprehensive_stats(db, date_from, date_to, anchor_id, anchor_name)

    salary_results = calc_anchor_salary(
        base_rate=base_rate, lead_commission=lead_commission,
        ad_commission=ad_commission, natural_commission=natural_commission,
        anchor_summary=stats["anchor_summary"]
    )

    return {
        "code": 0,
        "daily_detail": stats["daily_detail"],
        "daily_aggregated": stats["daily_aggregated"],
        "anchor_summary": salary_results,
        "overall_summary": stats["overall_summary"],
        "params": {
            "base_rate": base_rate, "lead_commission": lead_commission,
            "ad_commission": ad_commission, "natural_commission": natural_commission,
        },
    }


@router.get("/private-messages")
def admin_private_messages(date_from: str = None, date_to: str = None,
                           session_id: int = None, anchor_name: str = None,
                           has_lead: bool = None, page: int = 1, page_size: int = 50,
                           admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """后台私信分析查询API - 支持按日期/场次/主播/是否留资筛选"""
    from models import PrivateMessage, Session, Anchor, SessionAnchor

    query = db.query(
        PrivateMessage.id,
        PrivateMessage.session_id,
        PrivateMessage.nickname,
        PrivateMessage.douyin_id,
        PrivateMessage.has_lead,
        PrivateMessage.last_message_time,
        PrivateMessage.last_reply_time,
        PrivateMessage.pending_reply,
        PrivateMessage.message_count,
        PrivateMessage.ai_reply_count,
        PrivateMessage.created_at,
        Session.start_time.label('session_start'),
        func.group_concat(Anchor.name.distinct()).label('anchor_names'),
    ).join(Session, PrivateMessage.session_id == Session.id)\
     .outerjoin(SessionAnchor, Session.id == SessionAnchor.session_id)\
     .outerjoin(Anchor, SessionAnchor.anchor_id == Anchor.id)

    if date_from:
        query = query.filter(Session.start_time >= date_from)
    if date_to:
        query = query.filter(Session.start_time <= date_to + " 23:59:59")
    if session_id:
        query = query.filter(PrivateMessage.session_id == session_id)
    if has_lead is not None:
        query = query.filter(PrivateMessage.has_lead == has_lead)
    if anchor_name:
        query = query.filter(Anchor.name == anchor_name)

    # SQLite不支持nullslast，使用desc()即可
    query = query.group_by(PrivateMessage.id).order_by(PrivateMessage.last_message_time.desc())

    # 使用子查询方式获取总数（避免group_by后count的问题）
    total_sub = query.subquery()
    total = db.query(func.count()).select_from(total_sub).scalar() or 0
    items = query.offset((page - 1) * page_size).limit(page_size).all()

    # 统计摘要
    lead_count = db.query(func.count(PrivateMessage.id)).filter(
        PrivateMessage.has_lead == True
    )
    if session_id:
        lead_count = lead_count.filter(PrivateMessage.session_id == session_id)
    pm_lead = lead_count.scalar() or 0

    return {
        "code": 0,
        "data": [{
            "id": r.id, "session_id": r.session_id, "nickname": r.nickname,
            "douyin_id": r.douyin_id, "has_lead": r.has_lead,
            "last_message_time": r.last_message_time, "last_reply_time": r.last_reply_time,
            "pending_reply": r.pending_reply, "message_count": r.message_count,
            "ai_reply_count": r.ai_reply_count, "created_at": r.created_at,
            "session_start": r.session_start, "anchor_names": r.anchor_names or "",
        } for r in items],
        "total": total, "page": page, "page_size": page_size,
        "summary": {"has_lead_count": pm_lead},
    }


@router.post("/sessions/{session_id}/analyze")
def manual_analyze(session_id: int, admin=Depends(get_current_admin), db=Depends(get_db)):
    from services.ai_service import analyze_session
    try:
        report_id = analyze_session(db, session_id)
        return {"code": 0, "message": "分析完成", "report_id": report_id}
    except Exception as e:
        return {"code": 500, "message": str(e)}

class LeadUpdate(BaseModel):
    tags: Optional[str] = None
    is_valid: Optional[bool] = None
    notes: Optional[str] = None
    assigned_employee: Optional[str] = None

class LeadBatchUpdate(BaseModel):
    lead_ids: list[int]
    tags: Optional[str] = None
    is_valid: Optional[bool] = None

@router.put("/leads/{lead_id}")
def update_lead(lead_id: int, data: LeadUpdate, admin=Depends(get_current_admin), db=Depends(get_db)):
    lead = db.query(Lead).get(lead_id)
    if not lead:
        raise HTTPException(404, detail="线索不存在")
    if data.tags is not None:
        lead.tags = data.tags
    if data.is_valid is not None:
        lead.is_valid = data.is_valid
    if data.assigned_employee is not None:
        lead.assigned_employee = data.assigned_employee
    db.commit()
    return {"code": 0, "message": "线索已更新"}

@router.post("/leads/batch")
def batch_update_leads(data: LeadBatchUpdate, admin=Depends(get_current_admin), db=Depends(get_db)):
    leads = db.query(Lead).filter(Lead.id.in_(data.lead_ids)).all()
    for lead in leads:
        if data.tags is not None:
            lead.tags = data.tags
        if data.is_valid is not None:
            lead.is_valid = data.is_valid
    db.commit()
    return {"code": 0, "message": f"已更新 {len(leads)} 条线索"}

@router.get("/leads")
def list_leads(session_id: int = None, keyword: str = None, city: str = None, is_valid: str = None, source: str = None, page: int = 1, size: int = 20, admin=Depends(get_current_admin), db=Depends(get_db)):
    query = db.query(Lead).order_by(Lead.created_at.desc())
    if session_id:
        query = query.filter(Lead.session_id == session_id)
    if keyword:
        query = query.filter(Lead.nickname.contains(keyword))
    if city:
        query = query.filter(Lead.city == city)
    if is_valid:
        if is_valid == "true":
            query = query.filter(Lead.is_valid == True)
        elif is_valid == "false":
            query = query.filter(Lead.is_valid == False)
    if source:
        query = query.filter(Lead.source == source)
    total = query.count()
    items = query.offset((page-1)*size).limit(size).all()
    return {"code": 0, "data": {"items": [{"id": l.id, "session_id": l.session_id, "lead_time": l.lead_time, "nickname": l.nickname, "city": l.city, "path": l.path, "tags": l.tags, "is_valid": l.is_valid, "is_deal": l.is_deal, "source": l.source, "assigned_employee": l.assigned_employee, "ad_account": l.ad_account, "phone_masked": l.phone_masked, "product_name": l.product_name} for l in items], "total": total}}

# ===== 排班方案 API =====

@router.get("/schedule/plans")
def list_schedule_plans(active_only: bool = False, admin=Depends(get_current_admin), db=Depends(get_db)):
    query = db.query(SchedulePlan).order_by(SchedulePlan.created_at.desc())
    if active_only:
        query = query.filter(SchedulePlan.is_active == True)
    plans = query.all()
    return {"code": 0, "data": [{"id": p.id, "name": p.name, "start_time": p.start_time, "end_time": p.end_time,
        "time_granularity": p.time_granularity, "room_count": p.room_count, "anchor_count": p.anchor_count,
        "is_active": p.is_active, "notes": p.notes} for p in plans]}

@router.post("/schedule/plans")
def create_schedule_plan(data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    plan = SchedulePlan(
        name=data["name"], start_time=data["start_time"], end_time=data["end_time"],
        time_granularity=data.get("time_granularity", 60), room_count=data.get("room_count", 2),
        anchor_count=data.get("anchor_count", 5), is_active=data.get("is_active", True),
        notes=data.get("notes"))
    db.add(plan); db.commit()
    _generate_slots(db, plan)
    db.commit()
    return {"code": 0, "plan_id": plan.id}

@router.put("/schedule/plans/{plan_id}")
def update_schedule_plan(plan_id: int, data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    plan = db.query(SchedulePlan).get(plan_id)
    if not plan: raise HTTPException(404)
    rebuild_slots = False
    if "name" in data: plan.name = data["name"]
    if "start_time" in data: plan.start_time = data["start_time"]; rebuild_slots = True
    if "end_time" in data: plan.end_time = data["end_time"]; rebuild_slots = True
    if "time_granularity" in data: plan.time_granularity = data["time_granularity"]; rebuild_slots = True
    if "room_count" in data: plan.room_count = data["room_count"]; rebuild_slots = True
    if "anchor_count" in data: plan.anchor_count = data["anchor_count"]
    if "is_active" in data: plan.is_active = data["is_active"]
    if "notes" in data: plan.notes = data["notes"]
    if rebuild_slots:
        db.query(ScheduleSlot).filter(ScheduleSlot.plan_id == plan_id).delete()
        _generate_slots(db, plan)
    db.commit()
    return {"code": 0}

@router.delete("/schedule/plans/{plan_id}")
def delete_schedule_plan(plan_id: int, admin=Depends(get_current_admin), db=Depends(get_db)):
    binding_count = db.query(ScheduleBinding).filter(ScheduleBinding.plan_id == plan_id).count()
    if binding_count > 0:
        return {"code": 400, "message": f"该方案已被 {binding_count} 天绑定，请先解除绑定"}
    db.query(ScheduleSlot).filter(ScheduleSlot.plan_id == plan_id).delete()
    db.query(SchedulePlan).filter(SchedulePlan.id == plan_id).delete()
    db.commit()
    return {"code": 0}

@router.get("/schedule/plans/{plan_id}/slots")
def get_schedule_slots(plan_id: int, admin=Depends(get_current_admin), db=Depends(get_db)):
    plan = db.query(SchedulePlan).get(plan_id)
    if not plan: raise HTTPException(404)
    slots = db.query(ScheduleSlot).filter(ScheduleSlot.plan_id == plan_id).order_by(ScheduleSlot.time_slot, ScheduleSlot.room_index).all()
    return {"code": 0, "data": {
        "plan": {"id": plan.id, "name": plan.name, "start_time": plan.start_time, "end_time": plan.end_time,
            "time_granularity": plan.time_granularity, "room_count": plan.room_count, "anchor_count": plan.anchor_count,
            "is_active": plan.is_active, "notes": plan.notes},
        "slots": [{"id": s.id, "time_slot": s.time_slot, "room_index": s.room_index,
            "slot_status": s.slot_status, "anchor_slot": s.anchor_slot} for s in slots]}}

@router.put("/schedule/plans/{plan_id}/slots")
def save_schedule_slots(plan_id: int, data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    plan = db.query(SchedulePlan).get(plan_id)
    if not plan: raise HTTPException(404)
    db.query(ScheduleSlot).filter(ScheduleSlot.plan_id == plan_id).delete()
    for item in data.get("slots", []):
        slot = ScheduleSlot(plan_id=plan_id, time_slot=item["time_slot"], room_index=item["room_index"],
            slot_status=item.get("slot_status", "on_air_anchor"), anchor_slot=item.get("anchor_slot"))
        db.add(slot)
    db.commit()
    return {"code": 0}

# ===== 日期绑定 API =====

@router.get("/schedule/bindings")
def list_schedule_bindings(year: int = None, month: int = None, admin=Depends(get_current_admin), db=Depends(get_db)):
    query = db.query(ScheduleBinding)
    if year and month:
        prefix = f"{year}-{month:02d}"
        query = query.filter(ScheduleBinding.date.like(f"{prefix}%"))
    bindings = query.order_by(ScheduleBinding.date).all()
    return {"code": 0, "data": [{"id": b.id, "date": b.date, "plan_id": b.plan_id,
        "plan_name": b.plan.name if b.plan else None, "anchor_mapping": json.loads(b.anchor_mapping)} for b in bindings]}

@router.post("/schedule/bindings")
def create_schedule_binding(data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    dates = data.get("dates", [])
    plan_id = data.get("plan_id")
    anchor_mapping = data.get("anchor_mapping", {})
    if not dates or not plan_id:
        raise HTTPException(400, "缺少日期或方案ID")
    plan = db.query(SchedulePlan).get(plan_id)
    if not plan: raise HTTPException(404, "方案不存在")
    mapping_json = json.dumps(anchor_mapping, ensure_ascii=False)
    results = []
    for date_str in dates:
        existing = db.query(ScheduleBinding).filter(ScheduleBinding.date == date_str).first()
        if existing:
            existing.plan_id = plan_id
            existing.anchor_mapping = mapping_json
            results.append({"date": date_str, "action": "updated"})
        else:
            binding = ScheduleBinding(date=date_str, plan_id=plan_id, anchor_mapping=mapping_json)
            db.add(binding)
            results.append({"date": date_str, "action": "created"})
        _sync_session_anchors(db, date_str, plan, anchor_mapping)
    db.commit()
    return {"code": 0, "data": results}

@router.delete("/schedule/bindings/{date}")
def delete_schedule_binding(date: str, admin=Depends(get_current_admin), db=Depends(get_db)):
    binding = db.query(ScheduleBinding).filter(ScheduleBinding.date == date).first()
    if not binding: raise HTTPException(404)
    db.delete(binding); db.commit()
    return {"code": 0}

@router.put("/schedule/bindings/{binding_id}/slots")
def update_binding_slots(binding_id: int, data: dict, admin=Depends(get_current_admin), db=Depends(get_db)):
    """更新绑定的排班时段（支持临时加班添加额外时段）"""
    binding = db.query(ScheduleBinding).filter(ScheduleBinding.id == binding_id).first()
    if not binding: raise HTTPException(404, "排班绑定不存在")
    plan = db.query(SchedulePlan).get(binding.plan_id)
    if not plan: raise HTTPException(404, "排班方案不存在")
    slots_data = data.get("slots", [])
    # 删除旧slots，写入新slots
    db.query(ScheduleSlot).filter(ScheduleSlot.plan_id == plan.id).delete()
    for s in slots_data:
        slot = ScheduleSlot(
            plan_id=plan.id,
            time_slot=s.get("time_slot"),
            room_index=s.get("room_index", 1),
            slot_status=s.get("slot_status", "on_air_anchor"),
            anchor_slot=s.get("anchor_slot")
        )
        db.add(slot)
    db.commit()
    return {"code": 0, "message": "排班时段已更新"}

@router.get("/schedule/log")
def get_schedule_log(date: str, admin=Depends(get_current_admin), db=Depends(get_db)):
    """获取某日排班日志：方案详情+时段明细+主播映射"""
    binding = db.query(ScheduleBinding).filter(ScheduleBinding.date == date).first()
    if not binding:
        return {"code": 0, "data": None}
    plan = db.query(SchedulePlan).get(binding.plan_id)
    if not plan:
        return {"code": 0, "data": None}
    slots = db.query(ScheduleSlot).filter(ScheduleSlot.plan_id == plan.id).order_by(ScheduleSlot.time_slot, ScheduleSlot.room_index).all()
    anchor_mapping = json.loads(binding.anchor_mapping) if binding.anchor_mapping else {}
    anchor_ids = [int(v) for v in anchor_mapping.values() if v]
    anchors = db.query(Anchor).filter(Anchor.id.in_(anchor_ids)).all() if anchor_ids else []
    anchor_dict = {a.id: {"id": a.id, "name": a.name, "gender": a.gender, "style": a.style} for a in anchors}
    resolved_mapping = {}
    for slot_key, aid in anchor_mapping.items():
        if aid and int(aid) in anchor_dict:
            resolved_mapping[slot_key] = anchor_dict[int(aid)]
    sessions = db.query(Session).filter(Session.start_time.like(f"{date}%")).order_by(Session.start_time).all()
    return {"code": 0, "data": {
        "binding": {"id": binding.id, "date": binding.date, "plan_id": binding.plan_id},
        "plan": {"id": plan.id, "name": plan.name, "start_time": plan.start_time, "end_time": plan.end_time,
            "time_granularity": plan.time_granularity, "room_count": plan.room_count, "anchor_count": plan.anchor_count},
        "slots": [{"time_slot": s.time_slot, "room_index": s.room_index,
            "slot_status": s.slot_status, "anchor_slot": s.anchor_slot} for s in slots],
        "anchor_mapping": resolved_mapping,
        "sessions": [{"id": s.id, "start_time": s.start_time[:16], "end_time": s.end_time[:16] if s.end_time else None,
            "title": s.title} for s in sessions]}}

# ===== 直播账户 API =====

@router.get("/ad-accounts")
def list_ad_accounts(admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    accounts = db.query(AdAccount).order_by(AdAccount.created_at.desc()).all()
    return {"code": 0, "data": [{"id": a.id, "account_name": a.account_name, "merchant_id": a.merchant_id, "account_id": a.account_id, "notes": a.notes} for a in accounts]}

@router.post("/ad-accounts")
def create_ad_account(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    account = AdAccount(account_name=data["account_name"], merchant_id=data.get("merchant_id"), account_id=data.get("account_id"), notes=data.get("notes"))
    db.add(account); db.commit(); db.refresh(account)
    return {"code": 0, "account_id": account.id}

@router.put("/ad-accounts/{id}")
def update_ad_account(id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    account = db.query(AdAccount).get(id)
    if not account: raise HTTPException(404)
    if "account_name" in data: account.account_name = data["account_name"]
    if "merchant_id" in data: account.merchant_id = data["merchant_id"]
    if "account_id" in data: account.account_id = data["account_id"]
    if "notes" in data: account.notes = data["notes"]
    db.commit(); return {"code": 0}

@router.delete("/ad-accounts/{id}")
def delete_ad_account(id: int, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    binding_count = db.query(RoomAccountBinding).filter(RoomAccountBinding.ad_account_id == id).count()
    if binding_count > 0:
        return {"code": 400, "message": f"该账户已被 {binding_count} 个直播间绑定，请先解除绑定"}
    db.query(AdAccount).filter(AdAccount.id == id).delete()
    db.commit(); return {"code": 0}

# ===== 直播间账户绑定 API =====

@router.get("/room-bindings")
def list_room_bindings(date: str = None, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    query = db.query(RoomAccountBinding)
    if date:
        query = query.filter(RoomAccountBinding.date == date)
    bindings = query.order_by(RoomAccountBinding.date, RoomAccountBinding.room_index).all()
    return {"code": 0, "data": [{"id": b.id, "date": b.date, "room_index": b.room_index,
        "ad_account_id": b.ad_account_id, "account_name": b.ad_account.account_name if b.ad_account else None,
        "account_id": b.ad_account.account_id if b.ad_account else None} for b in bindings]}

@router.post("/room-bindings")
def save_room_bindings(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """批量保存某日的直播间账户绑定"""
    date = data.get("date")
    bindings = data.get("bindings", [])  # [{room_index: 1, ad_account_id: 3}, ...]
    if not date:
        raise HTTPException(400, "缺少日期")
    # 删除当日旧绑定
    db.query(RoomAccountBinding).filter(RoomAccountBinding.date == date).delete()
    for item in bindings:
        if item.get("ad_account_id"):
            binding = RoomAccountBinding(date=date, room_index=item["room_index"], ad_account_id=item["ad_account_id"])
            db.add(binding)
    db.commit()
    return {"code": 0}

@router.post("/room-bindings/swap")
def swap_room_accounts(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """互换两个直播间的账户绑定"""
    date = data.get("date")
    room_a = data.get("room_a")
    room_b = data.get("room_b")
    if not date or not room_a or not room_b:
        raise HTTPException(400, "缺少参数")
    binding_a = db.query(RoomAccountBinding).filter(RoomAccountBinding.date == date, RoomAccountBinding.room_index == room_a).first()
    binding_b = db.query(RoomAccountBinding).filter(RoomAccountBinding.date == date, RoomAccountBinding.room_index == room_b).first()
    account_a = binding_a.ad_account_id if binding_a else None
    account_b = binding_b.ad_account_id if binding_b else None
    if binding_a:
        binding_a.ad_account_id = account_b
    elif account_b:
        new_a = RoomAccountBinding(date=date, room_index=room_a, ad_account_id=account_b)
        db.add(new_a)
    if binding_b:
        binding_b.ad_account_id = account_a
    elif account_a:
        new_b = RoomAccountBinding(date=date, room_index=room_b, ad_account_id=account_a)
        db.add(new_b)
    db.commit()
    return {"code": 0}

# ===== 排班辅助函数 =====

def _parse_time_minutes(t: str) -> int:
    """将时间字符串转为分钟数，支持跨天（如 25:00 = 次日01:00）"""
    parts = t.replace("：", ":").split(":")
    return int(parts[0]) * 60 + int(parts[1])

def _generate_slots(db, plan: SchedulePlan):
    """根据方案大纲自动生成时段行，默认开播有主播并自动分配主播编号"""
    start_min = _parse_time_minutes(plan.start_time)
    end_min = _parse_time_minutes(plan.end_time)
    if end_min <= start_min:
        end_min += 24 * 60  # 跨天
    gran = plan.time_granularity
    current = start_min
    while current < end_min:
        slot_start_h = current // 60
        slot_start_m = current % 60
        next_time = min(current + gran, end_min)
        next_h = next_time // 60
        next_m = next_time % 60
        time_slot = f"{slot_start_h:02d}:{slot_start_m:02d}-{next_h:02d}:{next_m:02d}"
        for room_idx in range(1, plan.room_count + 1):
            # 自动分配主播编号：按房间序号循环分配
            anchor_slot = ((room_idx - 1) % plan.anchor_count) + 1 if plan.anchor_count > 0 else None
            slot = ScheduleSlot(plan_id=plan.id, time_slot=time_slot, room_index=room_idx,
                slot_status="on_air_anchor", anchor_slot=anchor_slot)
            db.add(slot)
        current = next_time

def _sync_session_anchors(db, date_str: str, plan: SchedulePlan, anchor_mapping: dict):
    """根据日期绑定自动同步 SessionAnchor 记录，从排班时段中提取 on_time/off_time/anchor_order"""
    sessions = db.query(Session).filter(Session.start_time.like(f"{date_str}%")).all()
    if not sessions:
        return

    # 获取该方案的所有时段，仅处理"开播有主播"的时段
    slots = db.query(ScheduleSlot).filter(
        ScheduleSlot.plan_id == plan.id,
        ScheduleSlot.slot_status == "on_air_anchor"
    ).all()

    # 构建每个锚点的时段列表: [{anchor_id, on_time, off_time, on_minutes, off_minutes}]
    seen = set()
    anchor_blocks = []
    for slot in slots:
        anchor_slot = slot.anchor_slot
        if anchor_slot is None:
            continue
        anchor_id = anchor_mapping.get(str(anchor_slot))
        if not anchor_id:
            continue
        anchor_id = int(anchor_id)
        parts = slot.time_slot.split("-")
        on_time = parts[0]
        off_time = parts[1]
        on_min = _parse_time_minutes(on_time)
        off_min = _parse_time_minutes(off_time)
        key = (anchor_id, on_min, off_min)
        if key in seen:
            continue
        seen.add(key)
        anchor_blocks.append({
            "anchor_id": anchor_id,
            "on_time": on_time,
            "off_time": off_time,
            "on_minutes": on_min,
            "off_minutes": off_min,
        })

    if not anchor_blocks:
        return

    # 按上播时间排序
    anchor_blocks.sort(key=lambda x: (x["on_minutes"], x["anchor_id"]))

    # 合并同一锚点的连续时段
    merged = []
    for block in anchor_blocks:
        if merged and merged[-1]["anchor_id"] == block["anchor_id"] and merged[-1]["off_time"] == block["on_time"]:
            merged[-1]["off_time"] = block["off_time"]
            merged[-1]["off_minutes"] = block["off_minutes"]
        else:
            merged.append(dict(block))

    # 同步到各场次：根据场次时间范围匹配对应的anchor_blocks
    for s in sessions:
        db.query(SessionAnchor).filter(SessionAnchor.session_id == s.id).delete()

        # 解析场次的时间范围
        session_start_time = s.start_time[11:16] if s.start_time and len(s.start_time) >= 16 else None
        session_end_time = s.end_time[11:16] if s.end_time and len(s.end_time) >= 16 else None
        if not session_start_time:
            continue

        session_start_min = _parse_time_minutes(session_start_time)
        session_end_min = _parse_time_minutes(session_end_time) if session_end_time else session_start_min + 360

        # 筛选与该场次时间重叠的anchor_blocks
        matched_blocks = []
        for block in merged:
            # 判断block与session是否有时间重叠
            block_start = block["on_minutes"]
            block_end = block["off_minutes"]
            # 跨天处理：如果off < on，说明跨天
            if block_end < block_start:
                block_end += 24 * 60
            if session_end_min < session_start_min:
                session_end_min += 24 * 60
            # 重叠条件：block_start < session_end && block_end > session_start
            if block_start < session_end_min and block_end > session_start_min:
                matched_blocks.append(block)

        for order, block in enumerate(matched_blocks, 1):
            db.add(SessionAnchor(
                session_id=s.id,
                anchor_id=block["anchor_id"],
                on_time=block["on_time"],
                off_time=block["off_time"],
                anchor_order=order
            ))

# ===== 数据库备份/导入导出 API =====

DB_PATH = Path(__file__).parent / "data.db"
BACKUP_DIR = Path(__file__).parent / "backups"

@router.get("/db/backup")
def backup_database(admin=Depends(get_current_admin)):
    """下载当前数据库完整备份"""
    if not DB_PATH.exists():
        raise HTTPException(404, "数据库文件不存在")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return FileResponse(str(DB_PATH), media_type="application/octet-stream",
        filename=f"alive_broadcast_{timestamp}.db")

@router.get("/db/backups")
def list_backups(admin=Depends(get_current_admin)):
    """列出自动备份文件"""
    BACKUP_DIR.mkdir(exist_ok=True)
    backups = sorted(BACKUP_DIR.glob("data_*.db"), reverse=True)
    return {"code": 0, "data": [{"name": b.name, "size_kb": round(b.stat().st_size/1024, 1),
        "modified": datetime.fromtimestamp(b.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")} for b in backups]}

@router.get("/db/backups/{filename}")
def download_backup(filename: str, admin=Depends(get_current_admin)):
    """下载指定备份文件"""
    filepath = BACKUP_DIR / filename
    if not filepath.exists() or not filepath.name.startswith("data_"):
        raise HTTPException(404, "备份文件不存在")
    return FileResponse(str(filepath), media_type="application/octet-stream", filename=filename)

@router.post("/db/upload-restore")
async def upload_restore_database(admin=Depends(get_current_admin), file: bytes = None):
    """上传数据库文件并恢复"""
    if not file:
        raise HTTPException(400, "未上传文件")
    # 先备份当前数据库
    BACKUP_DIR.mkdir(exist_ok=True)
    pre_restore_name = f"data_pre_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    if DB_PATH.exists():
        shutil.copy2(DB_PATH, BACKUP_DIR / pre_restore_name)
    # 写入新数据库
    with open(DB_PATH, 'wb') as f:
        f.write(file)
    return {"code": 0, "message": "数据库已恢复，请重启服务生效", "pre_restore_backup": pre_restore_name}

@router.post("/db/export-json")
def export_database_json(admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """导出全部数据为JSON格式"""
    from models import SessionMetric, Comment, HighIntentUser, PrivateMessage, Report
    data = {
        "export_time": datetime.now().isoformat(),
        "sessions": [{"id":s.id,"start_time":s.start_time,"end_time":s.end_time,"duration_minutes":s.duration_minutes,"analyzed":s.analyzed} for s in db.query(Session).all()],
        "session_metrics": [{"id":m.id,"session_id":m.session_id,"exposure_count":m.exposure_count,"cumulative_viewers":m.cumulative_viewers,"ad_spend":m.ad_spend,"total_leads":m.total_leads,"lead_cost":m.lead_cost,"max_online":m.max_online,"avg_online":m.avg_online} for m in db.query(SessionMetric).all()],
        "leads": [{"id":l.id,"session_id":l.session_id,"lead_time":l.lead_time,"nickname":l.nickname,"city":l.city,"path":l.path,"tags":l.tags,"ad_account":l.ad_account,"is_valid":l.is_valid,"is_deal":l.is_deal} for l in db.query(Lead).all()],
        "anchors": [{"id":a.id,"name":a.name,"gender":a.gender,"style":a.style} for a in db.query(Anchor).all()],
        "session_anchors": [{"id":sa.id,"session_id":sa.session_id,"anchor_id":sa.anchor_id} for sa in db.query(SessionAnchor).all()],
        "deals": [{"id":d.id,"session_id":d.session_id,"lead_id":d.lead_id,"customer_name":d.customer_name,"amount":d.amount,"deal_time":d.deal_time,"employee":d.employee} for d in db.query(Deal).all()],
        "schedule_plans": [{"id":p.id,"name":p.name,"start_time":p.start_time,"end_time":p.end_time,"time_granularity":p.time_granularity,"room_count":p.room_count,"anchor_count":p.anchor_count,"is_active":p.is_active,"notes":p.notes} for p in db.query(SchedulePlan).all()],
        "schedule_slots": [{"id":s.id,"plan_id":s.plan_id,"time_slot":s.time_slot,"room_index":s.room_index,"slot_status":s.slot_status,"anchor_slot":s.anchor_slot} for s in db.query(ScheduleSlot).all()],
        "schedule_bindings": [{"id":b.id,"date":b.date,"plan_id":b.plan_id,"anchor_mapping":b.anchor_mapping} for b in db.query(ScheduleBinding).all()],
        "settings": [{"id":s.id,"key":s.key,"value":s.value} for s in db.query(Setting).all()],
    }
    json_str = json.dumps(data, ensure_ascii=False, indent=2)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(io.BytesIO(json_str.encode('utf-8')), media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=alive_broadcast_{timestamp}.json"})

# ===== 报表Tab配置 API =====

@router.get("/dashboard-tabs")
def list_dashboard_tabs(admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    tabs = db.query(DashboardTab).order_by(DashboardTab.priority, DashboardTab.id).all()
    return {"code": 0, "data": [{"id": t.id, "name": t.name, "priority": t.priority,
        "is_system": t.is_system, "chart_type": t.chart_type, "metrics_config": json.loads(t.metrics_config) if t.metrics_config else [],
        "system_prompt": t.system_prompt, "user_prompt": t.user_prompt} for t in tabs]}

@router.post("/dashboard-tabs")
def create_dashboard_tab(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    tab = DashboardTab(
        name=data["name"],
        priority=data.get("priority", 1),
        is_system=False,
        chart_type=data.get("chart_type", "line"),
        metrics_config=json.dumps(data.get("metrics_config", []), ensure_ascii=False),
        system_prompt=data.get("system_prompt"),
        user_prompt=data.get("user_prompt"),
    )
    db.add(tab); db.commit()
    return {"code": 0, "tab_id": tab.id}

@router.put("/dashboard-tabs/{tab_id}")
def update_dashboard_tab(tab_id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    tab = db.query(DashboardTab).get(tab_id)
    if not tab: raise HTTPException(404)
    if tab.is_system and "name" in data and data["name"] != "总览":
        return {"code": 400, "message": "系统Tab不可修改名称"}
    if "name" in data: tab.name = data["name"]
    if "priority" in data: tab.priority = data["priority"]
    if "chart_type" in data: tab.chart_type = data["chart_type"]
    if "metrics_config" in data: tab.metrics_config = json.dumps(data["metrics_config"], ensure_ascii=False)
    if "system_prompt" in data: tab.system_prompt = data["system_prompt"]
    if "user_prompt" in data: tab.user_prompt = data["user_prompt"]
    db.commit(); return {"code": 0}

@router.delete("/dashboard-tabs/{tab_id}")
def delete_dashboard_tab(tab_id: int, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    tab = db.query(DashboardTab).get(tab_id)
    if not tab: raise HTTPException(404)
    if tab.is_system:
        return {"code": 400, "message": "系统Tab不可删除"}
    db.query(TabAnalysis).filter(TabAnalysis.tab_id == tab_id).delete()
    db.delete(tab); db.commit()
    return {"code": 0}

# ===== API线索账号配置 API =====

@router.get("/clue-configs")
def list_clue_configs(admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    configs = db.query(ClueConfig).order_by(ClueConfig.created_at.desc()).all()
    return {"code": 0, "data": [{"id": c.id, "account_name": c.account_name, "account_id": c.account_id,
            "ad_account_id": c.ad_account_id,
            "ad_account_name": c.ad_account.account_name if c.ad_account else None,
            "client_key": c.client_key[:4] + "****" if c.client_key else "",
            "is_active": c.is_active, "poll_interval_seconds": c.poll_interval_seconds,
            "decrypt_enabled": c.decrypt_enabled, "notes": c.notes} for c in configs]}

@router.post("/clue-configs")
def create_clue_config(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    # 若传了ad_account_id，从直播账户自动获取merchant_id作为account_id
    account_id = data.get("account_id")
    ad_account_id = data.get("ad_account_id")
    if ad_account_id and not account_id:
        ad_acct = db.query(AdAccount).get(ad_account_id)
        if ad_acct:
            account_id = ad_acct.merchant_id or ad_acct.account_id
    config = ClueConfig(
        account_name=data["account_name"], account_id=account_id,
        ad_account_id=ad_account_id or None,
        client_key=data["client_key"], client_secret=data["client_secret"],
        is_active=data.get("is_active", True),
        poll_interval_seconds=data.get("poll_interval_seconds", 30),
        decrypt_enabled=data.get("decrypt_enabled", True),
        notes=data.get("notes"))
    db.add(config); db.commit()
    try:
        from services.scheduler import refresh_clue_job
        refresh_clue_job()
    except Exception:
        pass
    return {"code": 0, "config_id": config.id}

@router.put("/clue-configs/{config_id}")
def update_clue_config(config_id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    config = db.query(ClueConfig).get(config_id)
    if not config: raise HTTPException(404)
    if "account_name" in data: config.account_name = data["account_name"]
    if "ad_account_id" in data: config.ad_account_id = data["ad_account_id"] or None
    if "account_id" in data: config.account_id = data["account_id"]
    if "client_key" in data: config.client_key = data["client_key"]
    if "client_secret" in data: config.client_secret = data["client_secret"]
    if "is_active" in data: config.is_active = data["is_active"]
    if "poll_interval_seconds" in data: config.poll_interval_seconds = data["poll_interval_seconds"]
    if "decrypt_enabled" in data: config.decrypt_enabled = data["decrypt_enabled"]
    if "notes" in data: config.notes = data["notes"]
    # 如果更新了ad_account_id，自动同步account_id
    if "ad_account_id" in data and data.get("ad_account_id"):
        ad_acct = db.query(AdAccount).get(data["ad_account_id"])
        if ad_acct and ad_acct.merchant_id:
            config.account_id = ad_acct.merchant_id
    db.commit()
    try:
        from services.scheduler import refresh_clue_job
        refresh_clue_job()
    except Exception:
        pass
    return {"code": 0}

@router.delete("/clue-configs/{config_id}")
def delete_clue_config(config_id: int, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    db.query(ClueConfig).filter(ClueConfig.id == config_id).delete()
    db.commit()
    try:
        from services.scheduler import refresh_clue_job
        refresh_clue_job()
    except Exception:
        pass
    return {"code": 0}

@router.post("/clue-configs/{config_id}/test")
def test_clue_config(config_id: int, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    config = db.query(ClueConfig).get(config_id)
    if not config: raise HTTPException(404)
    from services.clue_service import get_access_token
    token = get_access_token(config.client_key, config.client_secret)
    if token:
        return {"code": 0, "message": "连接成功，已获取access_token"}
    else:
        return {"code": 400, "message": "连接失败，请检查Client Key和Secret"}


@router.post("/clue-configs/poll-now")
def poll_clues_now(admin=Depends(get_current_admin)):
    """手动触发线索轮询"""
    import threading
    from services.clue_service import poll_all_clues
    thread = threading.Thread(target=poll_all_clues, kwargs={"triggered_by": "manual"}, daemon=True)
    thread.start()
    return {"code": 0, "message": "线索轮询已触发，请稍后查看采集日志"}


@router.post("/clue-configs/assign-history")
def assign_history_clues(admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """批量分配历史未分配线索（非当天的线索）"""
    from services.assign_service import assign_new_clues
    count = assign_new_clues(db, today_only=False)
    return {"code": 0, "message": f"已分配{count}条历史线索", "count": count}


@router.get("/scheduler/status")
def get_scheduler_status(admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """获取调度器状态（是否暂停）"""
    from models import Setting
    paused = db.query(Setting).filter(Setting.key == "scheduler_paused").first()
    return {"code": 0, "data": {"paused": (paused.value == "true") if paused else False}}


@router.post("/scheduler/toggle")
def toggle_scheduler(admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """切换调度器暂停/恢复状态"""
    from models import Setting
    paused = db.query(Setting).filter(Setting.key == "scheduler_paused").first()
    if not paused:
        paused = Setting(key="scheduler_paused", value="true")
        db.add(paused)
        db.commit()
        return {"code": 0, "message": "调度器已暂停", "data": {"paused": True}}
    else:
        new_val = "false" if paused.value == "true" else "true"
        paused.value = new_val
        db.commit()
        is_paused = new_val == "true"
        return {"code": 0, "message": "调度器已恢复" if not is_paused else "调度器已暂停", "data": {"paused": is_paused}}


@router.get("/push-config")
def get_push_config(admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """获取推送配置"""
    from models import ClueConfig
    configs = db.query(ClueConfig).filter(ClueConfig.is_active == True).all()
    # 取第一个活跃账号的配置作为全局配置
    if configs:
        c = configs[0]
        return {"code": 0, "data": {
            "push_enabled": c.push_enabled if c.push_enabled is not None else True,
            "push_time_range_days": c.push_time_range_days or 1,
        }}
    return {"code": 0, "data": {"push_enabled": True, "push_time_range_days": 1}}


@router.put("/push-config")
def update_push_config(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """更新推送配置（全局，应用到所有活跃账号）"""
    from models import ClueConfig
    configs = db.query(ClueConfig).filter(ClueConfig.is_active == True).all()
    if not configs:
        return {"code": 404, "message": "无活跃的线索账号"}
    push_enabled = data.get("push_enabled", True)
    push_time_range_days = max(1, min(30, data.get("push_time_range_days", 1)))
    for c in configs:
        c.push_enabled = push_enabled
        c.push_time_range_days = push_time_range_days
    db.commit()
    return {"code": 0, "message": "推送配置已更新", "data": {
        "push_enabled": push_enabled,
        "push_time_range_days": push_time_range_days,
    }}


@router.get("/poll-logs")
def list_poll_logs(page: int = 1, size: int = 20, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """获取采集日志列表"""
    from models import PollLog
    total = db.query(PollLog).count()
    logs = db.query(PollLog).order_by(PollLog.id.desc()).offset((page - 1) * size).limit(size).all()
    return {"code": 0, "data": {
        "total": total, "page": page, "size": size,
        "items": [{"id": l.id, "account_name": l.account_name, "account_id": l.account_id,
                   "status": l.status, "new_count": l.new_count, "decrypt_count": l.decrypt_count,
                   "total_count": l.total_count, "message": l.message,
                   "duration_ms": l.duration_ms, "triggered_by": l.triggered_by,
                   "created_at": l.created_at} for l in logs]
    }}


# ===== 评论管理 =====
@router.put("/comments/{comment_id}/consultation")
def toggle_comment_consultation(comment_id: int, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """切换评论的咨询相关性标记"""
    from models import Comment
    comment = db.query(Comment).get(comment_id)
    if not comment:
        return {"code": 404, "message": "评论不存在"}
    comment.is_consultation = not (comment.is_consultation if comment.is_consultation is not None else True)
    db.commit()
    return {"code": 0, "data": {"is_consultation": comment.is_consultation}}


# ===== 招生团队管理 =====
@router.get("/recruit-teams")
def list_recruit_teams(admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    teams = db.query(RecruitTeam).order_by(RecruitTeam.id).all()
    result = []
    for t in teams:
        employees = db.query(RecruitEmployee).filter(RecruitEmployee.team_id == t.id, RecruitEmployee.is_active == True).order_by(RecruitEmployee.sort_order).all()
        result.append({
            "id": t.id, "name": t.name, "password": t.password,
            "require_password": t.require_password,
            "dingtalk_webhook": t.dingtalk_webhook, "dingtalk_secret": t.dingtalk_secret,
            "is_active": t.is_active, "created_at": t.created_at,
            "employees": [{"id": e.id, "name": e.name, "phone": e.phone,
                          "dingtalk_phone": e.dingtalk_phone, "sort_order": e.sort_order,
                          "is_active": e.is_active} for e in employees]
        })
    return {"code": 0, "data": result}


@router.post("/recruit-teams")
def create_recruit_team(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    team = RecruitTeam(
        name=data["name"],
        password=data.get("password"),
        require_password=data.get("require_password", False),
        dingtalk_webhook=data.get("dingtalk_webhook"),
        dingtalk_secret=data.get("dingtalk_secret"),
    )
    db.add(team)
    db.commit()
    db.refresh(team)
    # 同步创建员工
    for emp in data.get("employees", []):
        db.add(RecruitEmployee(
            team_id=team.id, name=emp["name"], phone=emp.get("phone"),
            dingtalk_phone=emp.get("dingtalk_phone"), sort_order=emp.get("sort_order", 0),
        ))
    db.commit()
    return {"code": 0, "team_id": team.id}


@router.put("/recruit-teams/{team_id}")
def update_recruit_team(team_id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    team = db.query(RecruitTeam).get(team_id)
    if not team: raise HTTPException(404, detail="团队不存在")
    if "name" in data: team.name = data["name"]
    if "password" in data: team.password = data["password"]
    if "require_password" in data: team.require_password = data["require_password"]
    if "dingtalk_webhook" in data: team.dingtalk_webhook = data["dingtalk_webhook"]
    if "dingtalk_secret" in data: team.dingtalk_secret = data["dingtalk_secret"]
    if "is_active" in data: team.is_active = data["is_active"]
    # 同步更新员工
    if "employees" in data:
        # 删除旧员工
        db.query(RecruitEmployee).filter(RecruitEmployee.team_id == team_id).delete()
        for emp in data["employees"]:
            db.add(RecruitEmployee(
                team_id=team_id, name=emp["name"], phone=emp.get("phone"),
                dingtalk_phone=emp.get("dingtalk_phone"), sort_order=emp.get("sort_order", 0),
            ))
    db.commit()
    return {"code": 0, "message": "团队已更新"}


@router.delete("/recruit-teams/{team_id}")
def delete_recruit_team(team_id: int, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    team = db.query(RecruitTeam).get(team_id)
    if not team: raise HTTPException(404, detail="团队不存在")
    # 检查是否有未完成的分配
    active_assignments = db.query(ClueAssignment).filter(
        ClueAssignment.team_id == team_id, ClueAssignment.status == "unclaimed"
    ).count()
    if active_assignments > 0:
        return {"code": 400, "message": f"该团队还有{active_assignments}条未领取线索，请先处理"}
    db.query(RecruitEmployee).filter(RecruitEmployee.team_id == team_id).delete()
    db.delete(team)
    db.commit()
    return {"code": 0, "message": "团队已删除"}


# ===== 线索分配管理 =====
@router.get("/clue-assignments")
def list_clue_assignments(date: Optional[str] = None, team_id: Optional[int] = None,
                          status: Optional[str] = None, page: int = 1, size: int = 50,
                          admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    query = db.query(ClueAssignment).join(ApiClue)
    if date:
        query = query.filter(ApiClue.create_time_detail.like(f"{date}%"))
    if team_id:
        query = query.filter(ClueAssignment.team_id == team_id)
    if status:
        query = query.filter(ClueAssignment.status == status)
    total = query.count()
    items = query.order_by(ClueAssignment.assigned_at.desc()).offset((page - 1) * size).limit(size).all()
    result = []
    for a in items:
        clue = a.clue
        result.append({
            "id": a.id, "clue_id": a.clue_id, "team_id": a.team_id, "employee_id": a.employee_id,
            "assigned_at": a.assigned_at, "claimed_at": a.claimed_at, "status": a.status,
            "feedback": a.feedback, "remark": a.remark,
            "phone_decrypted": clue.phone_decrypted if clue else None,
            "phone_masked": clue.phone_masked if clue else None,
            "is_decrypted": clue.is_decrypted if clue else False,
            "nickname": clue.nickname if clue else None,
            "name": clue.name if clue else None,
            "lead_time": clue.lead_time if clue else None,
            "create_time_detail": clue.create_time_detail if clue else None,
            "anchor_names": clue.anchor_names if clue else None,
            "weixin": clue.weixin if clue else None,
            "weixin_manual": clue.weixin_manual if clue else None,
            "city_name": clue.city_name if clue else None,
            "effective_state": clue.effective_state if clue else None,
            "source": clue.clue_source if clue else None,
            "employee_name": a.employee.name if a.employee else None,
            "team_name": a.team.name if a.team else None,
        })
    return {"code": 0, "data": {"items": result, "total": total}}


@router.put("/clue-assignments/{assignment_id}")
def update_clue_assignment(assignment_id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    assignment = db.query(ClueAssignment).get(assignment_id)
    if not assignment: raise HTTPException(404, detail="分配记录不存在")
    if "feedback" in data: assignment.feedback = data["feedback"]
    if "remark" in data: assignment.remark = data["remark"]
    if "status" in data: assignment.status = data["status"]
    db.commit()
    return {"code": 0, "message": "分配记录已更新"}


@router.put("/clue-assignments/{assignment_id}/reassign")
def reassign_clue(assignment_id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    assignment = db.query(ClueAssignment).get(assignment_id)
    if not assignment: raise HTTPException(404, detail="分配记录不存在")
    employee_id = data.get("employee_id")
    if employee_id:
        emp = db.query(RecruitEmployee).get(employee_id)
        if not emp: raise HTTPException(404, detail="员工不存在")
        assignment.employee_id = employee_id
    team_id = data.get("team_id")
    if team_id:
        assignment.team_id = team_id
    assignment.assigned_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    assignment.status = "unclaimed"
    assignment.claimed_at = None
    db.commit()
    return {"code": 0, "message": "线索已重新分配"}


@router.post("/clue-assignments")
def create_assignment(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """手动分配线索给指定员工"""
    from services.assign_service import assign_to_employee

    clue_id = data.get("clue_id")
    team_id = data.get("team_id")
    employee_id = data.get("employee_id")

    if not all([clue_id, team_id, employee_id]):
        return {"code": 400, "message": "缺少必填参数"}

    assignment = assign_to_employee(clue_id, team_id, employee_id, db)
    if not assignment:
        return {"code": 400, "message": "分配失败，请检查线索/团队/员工状态"}

    return {"code": 0, "message": "分配成功", "assignment_id": assignment.id}


@router.put("/api-clues/{clue_id}/weixin")
def update_clue_weixin(clue_id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    clue = db.query(ApiClue).get(clue_id)
    if not clue: raise HTTPException(404, detail="线索不存在")
    clue.weixin_manual = data.get("weixin_manual")
    db.commit()
    return {"code": 0, "message": "微信已更新"}


# ===== 手动添加私信线索 =====
@router.post("/manual-clue")
def create_manual_clue(req: ManualClueCreate, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """
    手动添加私信线索
    - 如果提供session_id，直接关联该场次
    - 如果提供message_time，自动匹配该时间所属的场次
    - 如果都未提供，session_id为null，后续可手动关联
    """
    from sqlalchemy import and_
    import time, random
    
    # 生成线索ID: manual_{timestamp}_{random4}
    timestamp = int(time.time())
    random4 = f"{random.randint(1000, 9999)}"
    clue_id_str = f"manual_{timestamp}_{random4}"
    
    # 确定关联场次
    target_session_id = req.session_id
    
    if not target_session_id and req.message_time:
        # 根据私信时间自动匹配场次
        # 查找该时间所属的场次：start_time <= message_time <= end_time
        sessions = db.query(Session).filter(
            and_(
                Session.start_time <= req.message_time,
                Session.end_time >= req.message_time
            )
        ).first()
        if sessions:
            target_session_id = sessions.id
    
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 创建ApiClue记录
    new_clue = ApiClue(
        clue_id=clue_id_str,
        name=req.name,
        phone_decrypted=req.phone,
        weixin_manual=req.weixin,
        session_id=target_session_id,
        clue_source=req.source or "manual",
        raw_data=json.dumps({
            "manual_input": True,
            "notes": req.notes,
            "admin_name": admin if isinstance(admin, str) else str(getattr(admin, 'username', str(admin)))
        }, ensure_ascii=False),
        created_at=now_str,
    )
    db.add(new_clue)
    
    # 同步创建Lead记录（用于admin线索列表展示和管理）
    new_lead = Lead(
        session_id=target_session_id,
        lead_time=now_str[:16],
        nickname=req.name or "",
        city="",
        path="手动添加",
        tags="",
        is_valid=None,
        is_deal=False,
        source=req.source or "manual",
        assigned_employee=None,
        ad_account="",
    )
    db.add(new_lead)
    db.commit()
    db.refresh(new_clue)

    # 采集与推送已解耦：手动添加线索只入库，推送由调度器独立管理

    return {
        "code": 0,
        "message": "线索添加成功",
        "clue_id": new_clue.id,
        "session_id": target_session_id,
    }


@router.get("/api/sessions/match")
def match_session_by_time(message_time: str, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """
    根据私信时间自动匹配所属场次
    查询条件：start_time <= message_time <= end_time
    """
    from sqlalchemy import and_
    
    sessions = db.query(Session).filter(
        and_(
            Session.start_time <= message_time,
            Session.end_time >= message_time
        )
    ).order_by(Session.start_time.desc()).all()
    
    result = []
    for s in sessions:
        result.append({
            "id": s.id,
            "title": s.title,
            "start_time": s.start_time,
            "end_time": s.end_time,
        })
    
    return {"code": 0, "data": result}


# ===== 场次结束时间调整 =====
@router.put("/api/sessions/{session_id}/end-time")
def update_session_end_time(session_id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """
    修改场次结束时间
    - 接收新的end_time
    - 自动重算直播时长
    """
    session = db.query(Session).get(session_id)
    if not session:
        raise HTTPException(404, detail="场次不存在")
    
    new_end_time = data.get("end_time")
    if not new_end_time:
        raise HTTPException(400, detail="结束时间不能为空")
    
    # 更新结束时间
    session.end_time = new_end_time
    
    # 自动重算直播时长
    try:
        from datetime import datetime
        start = datetime.fromisoformat(session.start_time.replace("Z", "+00:00")) if "T" in session.start_time else datetime.strptime(session.start_time, "%Y-%m-%d %H:%M:%S")
        end = datetime.fromisoformat(new_end_time.replace("Z", "+00:00")) if "T" in new_end_time else datetime.strptime(new_end_time, "%Y-%m-%d %H:%M:%S")
        duration_minutes = int((end - start).total_seconds() / 60)
        # 更新session_metrics中的直播时长
        from models import SessionMetric
        metric = db.query(SessionMetric).filter(SessionMetric.session_id == session_id).first()
        if metric:
            metric.live_duration = duration_minutes
    except Exception as e:
        # 时间解析失败，不阻塞主流程
        pass
    
    db.commit()
    
    return {
        "code": 0,
        "message": "结束时间已更新",
        "end_time": new_end_time,
    }


# ===== Excel批量导入线索 =====
@router.post("/api/clues/excel-upload")
async def upload_excel_clues(file: bytes, session_id: Optional[int] = None, 
                             admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """
    Excel批量导入线索
    - 支持 .xlsx 和 .xls 格式
    - 必需列：手机号 或 姓名
    - 可选列：微信号、私信时间、备注
    """
    import io
    from openpyxl import load_workbook
    
    try:
        # 读取Excel文件
        wb = load_workbook(io.BytesIO(file), read_only=True)
        ws = wb.active
        
        # 获取表头
        headers = [cell.value for cell in ws[1] if cell.value]
        if not headers:
            return {"code": 400, "message": "Excel文件为空或格式错误"}
        
        # 检查必需列
        header_map = {h.lower().strip(): i for i, h in enumerate(headers)}
        phone_col = header_map.get("手机号") or header_map.get("phone") or header_map.get("电话")
        name_col = header_map.get("姓名") or header_map.get("name") or header_map.get("昵称") or header_map.get("nickname")
        weixin_col = header_map.get("微信号") or header_map.get("weixin") or header_map.get("微信")
        time_col = header_map.get("私信时间") or header_map.get("message_time") or header_map.get("留资时间") or header_map.get("lead_time")
        notes_col = header_map.get("备注") or header_map.get("notes")
        
        if phone_col is None and name_col is None:
            return {"code": 400, "message": "Excel必须包含'手机号'或'姓名'列"}
        
        # 解析数据
        import time, random
        timestamp = int(time.time())
        results = []
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # 跳过空行
                continue
            
            phone = str(row[phone_col]).strip() if phone_col is not None and row[phone_col] else None
            name = str(row[name_col]).strip() if name_col is not None and row[name_col] else None
            weixin = str(row[weixin_col]).strip() if weixin_col is not None and row[weixin_col] else None
            message_time = str(row[time_col]).strip() if time_col is not None and row[time_col] else None
            notes = str(row[notes_col]).strip() if notes_col is not None and row[notes_col] else None
            
            # 数据校验
            if not phone and not name:
                errors.append(f"第{row_idx}行：手机号和姓名至少填写一项")
                continue
            
            # 格式化手机号
            if phone:
                phone = phone.replace("-", "").replace(" ", "")
                if not phone.isdigit() or len(phone) < 11:
                    errors.append(f"第{row_idx}行：手机号格式错误 ({phone})")
                    continue
            
            # 自动匹配场次
            target_session_id = session_id
            if not target_session_id and message_time:
                try:
                    from sqlalchemy import and_
                    matched = db.query(Session).filter(
                        and_(
                            Session.start_time <= message_time,
                            Session.end_time >= message_time
                        )
                    ).first()
                    if matched:
                        target_session_id = matched.id
                except:
                    pass
            
            # 生成线索ID
            random4 = f"{random.randint(1000, 9999)}"
            clue_id_str = f"excel_{timestamp}_{random4}"
            
            results.append({
                "clue_id": clue_id_str,
                "name": name,
                "phone": phone,
                "weixin": weixin,
                "message_time": message_time,
                "session_id": target_session_id,
                "notes": notes,
            })
        
        wb.close()
        
        if errors:
            return {"code": 400, "message": "数据校验失败", "errors": errors}
        
        if not results:
            return {"code": 400, "message": "没有有效的数据行"}
        
        return {"code": 0, "data": results, "total": len(results)}
        
    except Exception as e:
        return {"code": 500, "message": f"文件解析失败: {str(e)}"}


@router.post("/api/clues/excel-confirm")
def confirm_excel_import(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    """
    确认Excel导入
    - 接收解析后的数据列表
    - 批量创建ApiClue记录
    """
    import time, random
    
    items = data.get("items", [])
    if not items:
        return {"code": 400, "message": "没有可导入的数据"}
    
    timestamp = int(time.time())
    created_count = 0
    
    for item in items:
        clue_id = item.get("clue_id") or f"excel_{timestamp}_{random.randint(1000, 9999)}"
        
        new_clue = ApiClue(
            clue_id=clue_id,
            name=item.get("name"),
            phone_decrypted=item.get("phone"),
            weixin_manual=item.get("weixin"),
            session_id=item.get("session_id"),
            clue_source="excel",  # 标记为Excel导入
            raw_data=json.dumps({
                "import_notes": item.get("notes"),
                "admin_name": admin if isinstance(admin, str) else str(getattr(admin, 'username', str(admin)))
            }, ensure_ascii=False),
            created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )
        db.add(new_clue)
        created_count += 1
    
    db.commit()
    
    return {"code": 0, "message": f"成功导入{created_count}条线索"}


# ===== 投流计划管理 =====
@router.get("/ad-plans")
def list_ad_plans(account_id: Optional[int] = None, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    from sqlalchemy import func
    query = db.query(AdPlan)
    if account_id:
        query = query.filter(AdPlan.account_id == account_id)
    plans = query.order_by(AdPlan.id.desc()).all()
    result = []
    for p in plans:
        total_spend = db.query(func.sum(AdPlanSpend.spend_amount)).filter(AdPlanSpend.plan_id == p.id).scalar() or 0
        total_result = db.query(func.sum(AdPlanSpend.result_count)).filter(AdPlanSpend.plan_id == p.id).scalar() or 0
        result.append({
            "id": p.id, "account_id": p.account_id, "plan_name": p.plan_name,
            "plan_id": p.plan_id, "bid_price": p.bid_price, "status": p.status,
            "start_date": p.start_date, "end_date": p.end_date, "notes": p.notes,
            "account_name": p.account.account_name if p.account else None,
            "total_spend": float(total_spend), "total_result": int(total_result),
        })
    return {"code": 0, "data": result}

@router.post("/ad-plans")
def create_ad_plan(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    plan = AdPlan(
        account_id=data["account_id"], plan_name=data["plan_name"],
        plan_id=data.get("plan_id"), bid_price=data.get("bid_price"),
        status=data.get("status", "active"), start_date=data.get("start_date"),
        end_date=data.get("end_date"), notes=data.get("notes"),
    )
    db.add(plan)
    db.commit()
    db.refresh(plan)
    return {"code": 0, "plan_id": plan.id}

@router.put("/ad-plans/{plan_id}")
def update_ad_plan(plan_id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    plan = db.query(AdPlan).get(plan_id)
    if not plan: raise HTTPException(404, detail="计划不存在")
    for k in ["plan_name", "plan_id", "bid_price", "status", "start_date", "end_date", "notes", "account_id"]:
        if k in data: setattr(plan, k, data[k])
    db.commit()
    return {"code": 0, "message": "计划已更新"}

@router.delete("/ad-plans/{plan_id}")
def delete_ad_plan(plan_id: int, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    plan = db.query(AdPlan).get(plan_id)
    if not plan: raise HTTPException(404, detail="计划不存在")
    db.query(AdPlanSpend).filter(AdPlanSpend.plan_id == plan_id).delete()
    db.delete(plan)
    db.commit()
    return {"code": 0, "message": "计划已删除"}


# ===== 投流花费记录 =====
@router.get("/ad-plan-spends")
def list_ad_plan_spends(plan_id: Optional[int] = None, record_date: Optional[str] = None,
                        page: int = 1, size: int = 50, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    query = db.query(AdPlanSpend)
    if plan_id: query = query.filter(AdPlanSpend.plan_id == plan_id)
    if record_date: query = query.filter(AdPlanSpend.record_date == record_date)
    total = query.count()
    items = query.order_by(AdPlanSpend.record_date.desc()).offset((page-1)*size).limit(size).all()
    result = [{
        "id": s.id, "plan_id": s.plan_id, "session_id": s.session_id,
        "spend_amount": s.spend_amount, "result_count": s.result_count,
        "record_date": s.record_date, "notes": s.notes,
        "plan_name": s.plan.plan_name if s.plan else None,
    } for s in items]
    return {"code": 0, "data": {"items": result, "total": total}}

@router.post("/ad-plan-spends")
def create_ad_plan_spend(data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    spend = AdPlanSpend(
        plan_id=data["plan_id"], session_id=data.get("session_id"),
        spend_amount=data.get("spend_amount", 0), result_count=data.get("result_count", 0),
        record_date=data["record_date"], notes=data.get("notes"),
    )
    db.add(spend)
    db.commit()
    return {"code": 0, "spend_id": spend.id}

@router.put("/ad-plan-spends/{spend_id}")
def update_ad_plan_spend(spend_id: int, data: dict, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    spend = db.query(AdPlanSpend).get(spend_id)
    if not spend: raise HTTPException(404, detail="记录不存在")
    for k in ["spend_amount", "result_count", "session_id", "record_date", "notes"]:
        if k in data: setattr(spend, k, data[k])
    db.commit()
    return {"code": 0, "message": "记录已更新"}

@router.delete("/ad-plan-spends/{spend_id}")
def delete_ad_plan_spend(spend_id: int, admin=Depends(get_current_admin), db: DBSession = Depends(get_db)):
    spend = db.query(AdPlanSpend).get(spend_id)
    if not spend: raise HTTPException(404, detail="记录不存在")
    db.delete(spend)
    db.commit()
    return {"code": 0, "message": "记录已删除"}
